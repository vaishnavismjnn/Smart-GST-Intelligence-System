# --- file: utils/excel_builder.py ---
# ═══════════════════════════════════════════════════════════════════════════
# EXCEL BUILDER
# Builds a styled two-sheet .xlsx workbook from a list of invoice records.
# Used by records.py (filtered or full export).
#
# Sheet 1 — "Invoices": every record in the input list, one row per invoice,
#   with colour-coded rows (red tint for invalid/mismatched, alternating
#   white/teal for clean records).
#
# Sheet 2 — "Summary": aggregated financials (deduplicated processed only)
#   including ITC claimable amount and recovery rate.
# ═══════════════════════════════════════════════════════════════════════════

import io
from utils.formatters import fmt_date, short_id
from utils.cleaner import clean_amount, deduplicate_records, get_valid_processed


def _build_excel(records: list, title: str = "GST Invoice Export") -> bytes:
    """
    Build a styled .xlsx from a list of record dicts.
    Returns raw bytes for st.download_button().

    WHY clean_amount() ON MONETARY COLUMNS:
      Raw MongoDB records can have amounts as OCR strings ("1,23,000").
      openpyxl writes whatever Python value it receives. If we write a string,
      Excel treats it as text — number_format doesn't apply and SUM formulas
      break. clean_amount() ensures every monetary cell is a Python float.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = Workbook()

    # ── Style constants ────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="00A896")
    ALT_FILL   = PatternFill("solid", fgColor="F0FAFA")
    ERR_FILL   = PatternFill("solid", fgColor="FDECEA")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

    HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    BODY_FONT  = Font(name="Arial", size=9)
    MONO_FONT  = Font(name="Courier New", size=9)
    TITLE_FONT = Font(name="Arial", bold=True, size=12, color="00A896")

    thin   = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")
    RIGHT  = Alignment(horizontal="right",  vertical="center")

    # ── Sheet 1: Invoices ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Invoices"

    # Title row (merged across all 11 columns)
    ws.merge_cells("A1:K1")
    ws["A1"] = title
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 24

    # Subtitle: record count
    ws.merge_cells("A2:K2")
    ws["A2"] = f"Total Records: {len(records)}"
    ws["A2"].font      = Font(name="Arial", size=9, color="5A5A5A", italic=True)
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 16

    # Header row (row 3)
    HEADERS = [
        ("ID",           8),  ("Merchant",    22), ("GSTIN",       18),
        ("Invoice No.", 14),  ("Date",        12), ("Total (₹)",   14),
        ("Taxable (₹)", 14),  ("GST (₹)",    13), ("GSTIN Valid", 11),
        ("Amt Match",   10),  ("Status",      10),
    ]
    for col_idx, (hdr, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=hdr)
        cell.fill      = HDR_FILL
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 20
    ws.freeze_panes = "A4"

    # Data rows
    for row_idx, r in enumerate(records, start=4):
        v       = r.get("validation", {}) or {}
        gst_ok  = v.get("gst_valid",    False)
        amt_ok  = v.get("amounts_match", False)
        is_warn = not gst_ok or not amt_ok

        row_fill = ERR_FILL if is_warn else (ALT_FILL if row_idx % 2 == 0 else WHITE_FILL)

        # All monetary values are passed through clean_amount() so openpyxl
        # always writes a float — number_format and SUM formulas work correctly.
        values = [
            short_id(r.get("_id", "")),
            r.get("MERCHANT")   or "—",
            r.get("GSTIN")      or "—",
            r.get("INVOICE_NO") or "—",
            fmt_date(r.get("INVOICE_DATE")),
            clean_amount(r.get("TOTAL_AMOUNT")),
            clean_amount(r.get("TAXABLE_AMOUNT")),
            clean_amount(r.get("GST_AMOUNT")),
            "✓ Valid"    if gst_ok else "✗ Invalid",
            "✓ Match"    if amt_ok else "✗ Mismatch",
            r.get("status") or "—",
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.fill   = row_fill
            if col_idx == 1:
                cell.font = MONO_FONT; cell.alignment = CENTER
            elif col_idx in (3, 4):
                cell.font = MONO_FONT; cell.alignment = LEFT
            elif col_idx in (6, 7, 8):
                # Monetary columns: always float after clean_amount,
                # so number_format applies unconditionally.
                cell.font          = MONO_FONT
                cell.alignment     = RIGHT
                cell.number_format = "₹#,##0.00"
            elif col_idx == 9:
                cell.font      = Font(name="Arial", size=9,
                                      color="007A6E" if gst_ok else "C0392B", bold=True)
                cell.alignment = CENTER
            elif col_idx == 10:
                cell.font      = Font(name="Arial", size=9,
                                      color="007A6E" if amt_ok else "C0392B", bold=True)
                cell.alignment = CENTER
            else:
                cell.font = BODY_FONT; cell.alignment = LEFT

    # ── Sheet 2: Summary ──────────────────────────────────────────────────────
    # WHY DEDUP HERE:
    #   The Summary sheet aggregates totals. If the user exports all records
    #   and some are duplicates, the summary would double-count them.
    #   deduplicate_records() ensures the summary reflects unique invoices only.
    #
    # WHY get_valid_processed FOR ITC:
    #   ITC can only be claimed from fully valid, processed, non-duplicate
    #   invoices with positive GST amounts. Using all processed records would
    #   overstate ITC by including invalid GSTINs and mismatched amounts.
    ws2 = wb.create_sheet(title="Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    deduped    = deduplicate_records(records)
    processed  = [r for r in deduped if r.get("status") == "processed"]
    valid      = get_valid_processed(records)   # same dedup + full 5-gate filter

    total_val  = sum(clean_amount(r.get("TOTAL_AMOUNT"))   for r in processed)
    total_gst  = sum(clean_amount(r.get("GST_AMOUNT"))     for r in processed)
    total_tax  = sum(clean_amount(r.get("TAXABLE_AMOUNT")) for r in processed)
    valid_cnt  = len(valid)
    itc_amt    = sum(
        clean_amount(r.get("GST_AMOUNT")) for r in valid
        if clean_amount(r.get("GST_AMOUNT")) > 0
    )

    ws2["A1"] = "GST Intelligence Platform — Export Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="00A896")
    ws2.merge_cells("A1:B1")
    ws2["A1"].alignment = CENTER

    itc_rate = f"{int(itc_amt / total_gst * 100) if total_gst else 0}%"

    summary_rows = [
        ("", ""),
        ("Metric", "Value"),
        ("Total Records (Export)",        len(deduped)),
        ("Processed Invoices",            len(processed)),
        ("Total Invoice Value",           total_val),
        ("Total Taxable Amount",          total_tax),
        ("Total GST Amount",              total_gst),
        ("Fully Valid (GSTIN + Amt ✓)",   f"{valid_cnt} / {len(processed)}"),
        ("ITC Claimable (Eligible GST)",  itc_amt),
        ("ITC Recovery Rate",             itc_rate),
    ]

    for r_idx, (label, val) in enumerate(summary_rows, start=2):
        lc = ws2.cell(row=r_idx, column=1, value=label)
        vc = ws2.cell(row=r_idx, column=2, value=val)
        if label == "Metric":
            for c in (lc, vc):
                c.fill = HDR_FILL; c.font = HDR_FONT
                c.border = BORDER; c.alignment = CENTER
        elif label:
            lc.font = Font(name="Arial", size=10, bold=True)
            lc.border = BORDER; lc.alignment = LEFT
            vc.font = MONO_FONT; vc.border = BORDER; vc.alignment = RIGHT
            if isinstance(val, float):
                vc.number_format = "₹#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()